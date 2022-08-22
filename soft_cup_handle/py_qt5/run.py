import os
import sys
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtCore import pyqtSignal, QObject
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtGui import QTextCursor
from matplotlib import pyplot as plt
from numpy import array, mat
from pylab import mpl
import UI_lan
from ToolsPackage import splitThread
from openpyxl import load_workbook

from fuction_1_extra import founction_1, handle_excel_2, founction_3
from python import k_means, excel_list_read_function_5, excel_list_read_function_4
from utils import assign_style_qt, get_merge_cell_list
import qdarkstyle


class Stream(QObject):
    """Redirects console output to text widget."""
    newText = pyqtSignal(str)

    def write(self, text):
        QtWidgets.QApplication.processEvents()
        self.newText.emit(str(text))


class anaxcelhandler(QtWidgets.QMainWindow, UI_lan.Ui_MainWindow):

    def __init__(self, parent=None):
        super(anaxcelhandler, self).__init__(parent)
        if getattr(sys, 'frozen', False):
            self.frozen = 'ever so'
            self.bundle_dir = sys._MEIPASS
        else:
            self.bundle_dir = os.path.dirname(os.path.abspath(__file__))
        self.setupUi(self)
        # self.setWindowIcon(QtGui.QIcon(self.bundle_dir + '/icons/icon.png'))
        # self.setStyleSheet(open("Dark/darkstyle.qss", "r").read())
        # self.setStyleSheet(open("qss/1.qss", "r").read())

        '''
        此处为多个槽函数的连接使用
        '''

        self.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.pushButtonbrowse.clicked.connect(self.openFileNamesDialog)
        self.pushButtonclear.clicked.connect(self.clearwidget)

        self.pushButtonload.clicked.connect(self.LoadProcess)
        # self.pushButtonsplit.clicked.connect(self.SplitProcess)
        self.pushButtonsecond.clicked.connect(self.second_function)
        self.pushButtonfirst.clicked.connect(self.first_function)
        self.pushButtonthird.clicked.connect(self.third_function)
        self.pushButtonforth.clicked.connect(self.forth_function)
        self.pushButtonfifth.clicked.connect(self.fifth_function)
        self.statusbar.showMessage('电力客户行为分析小程序')  # 界面左下角的文本
        self.comboBoxfiletype.addItems(['files'])

        # ==========log=====
        sys.stdout = Stream(newText=self.onUpdateText)
        # ==========log=====

        # ==========show====
        self.flag_confirm = False
        self.activate_file = [None, None]
        self.comboBox_wb.activated.connect(self.wbActivated)
        self.comboBox_ws.activated.connect(self.wsActivated)
        self.tableWidget.itemClicked.connect(self.handleItemClick)
        self.pushButton_clear_idx.clicked.connect(self.clear_idx)
        self.pushButton_confirm_idx.clicked.connect(self.confirm_idx)

        # ==========show====

        # ==========context===
        self.infos = {}
        self.infos_bak = {}

        print("此工具功能一为选择一个第一行包含用户编号、缴费日期、缴费金额的excel电力用户表")
        print("经过处理后，将会得到以编号为主的每个用户的平均缴费金额以及平均缴费次数")
        print("------------------------------------")
        print("此工具功能二为选择经过功能一处理后得到的excel表")
        print("经过处理后，将会对每个用户的电费缴费情况根据四种客户类型进行归类")
        print("------------------------------------")
        print("此工具功能三为选择一个第一行包含用户编号、缴费日期、缴费金额的excel电力用户表")
        print("经过处理后，将会根据时间序列，预测得到最有可能成为高价值客户的TOP5")
        print("------------------------------------")
        print("此工具功能四为选择经过功能一处理后得到的excel表")
        print("经过处理后，会得到一个根据聚类算法得到的聚类结果")
        print("------------------------------------")
        print("此工具功能五为选择一个第一行包含小时和电量的小时用电量excel表")
        print("经过处理后，将会得到一个小时和电量的折线图")
        print("-----------------------------------")

    def use_palette(self):
        self.setWindowTitle("设置背景图片")
        window_pale = QtGui.QPalette()
        window_pale.setBrush(self.backgroundRole(), QtGui.QBrush(QtGui.QPixmap(self.bundle_dir + '/icons/bg.jpeg')))
        self.setPalette(window_pale)

    def onUpdateText(self, text):
        """Write console output to text widget."""
        cursor = self.textBrowserlog.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(text)
        self.textBrowserlog.setTextCursor(cursor)
        self.textBrowserlog.ensureCursorVisible()

    def openFileNamesDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filterxls = "XLS (*.xls *.XLS)"
        filterxlsx = "XLSX (*.xlsx *.XLSX)"
        filterpng = "PNG (*.png *.PNG)"
        filterjpg = "JPG (*.jpg *.JPG *.png *.PNG *.xlsx *.XLSX)"
        filterfiles = "XLSX (*.xlsx *.XLSX *.jpg *.JPG *.png *.PNG *.xls *.XLS)"
        print('打开文件')
        if self.comboBoxfiletype.currentIndex() == 0:
            files, _ = QFileDialog.getOpenFileNames(self, "Select  Files", filter=filterfiles, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(file)

    def clearwidget(self):
        self.listWidget.clear()
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_wb.clear()
        self.comboBox_ws.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def clearcontext_all(self):
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_wb.clear()
        self.comboBox_ws.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def clearcontext_show(self):
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)

    def clear_idx(self):
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def assign_dict(self, dict1, dict2):
        for k, v in dict1.items():
            if isinstance(v, dict):
                dict_tmp = dict()
                dict2[k] = self.assign_dict(v, dict_tmp)
            else:
                dict2[k] = v
        return dict2

    def confirm_idx(self):
        self.infos_bak = self.assign_dict(self.infos, self.infos_bak)

        x = self.comboBox_x.itemText(self.comboBox_x.currentIndex())
        y = self.comboBox_y.itemText(self.comboBox_y.currentIndex())

        r1 = self.comboBox_r1.itemText(self.comboBox_r1.currentIndex())
        r2 = self.comboBox_r2.itemText(self.comboBox_r2.currentIndex())

        wb = self.comboBox_wb.itemText(self.comboBox_wb.currentIndex())
        ws = self.comboBox_ws.itemText(self.comboBox_ws.currentIndex())

        if wb == '' or ws == '':
            QMessageBox.about(self, "hi,兰神", '先load文件')
        else:
            x = int(x) if x != '' else x
            y = int(y) if y != '' else y
            r1 = int(r1) if r1 != '' else r1
            r2 = int(r2) if r2 != '' else r2

            if self.checkBox_book.isChecked():
                print('book')
                key_idx = [x, y]
                rg = [r1, 'last']
                for wb_k in self.infos_bak.keys():
                    ws_keys = self.infos_bak[wb_k]['sheet_names']
                    for ws_k in ws_keys.keys():
                        self.infos_bak[wb_k]['sheet_names'][ws_k] = [key_idx, rg]
            elif self.checkBox_sheet.isChecked():
                print('sheet')
                key_idx = [x, y]
                rg = [r1, 'last']
                ws_keys = self.infos_bak[wb]['sheet_names']
                for ws_k in ws_keys.keys():
                    self.infos_bak[wb]['sheet_names'][ws_k] = [key_idx, rg]
            else:
                print('cell')
                key_idx = [x, y]
                rg = [r1, r2]
                self.infos_bak[wb]['sheet_names'][ws] = [key_idx, rg]
            self.flag_confirm = True

    def selectall(self):
        self.listWidget.selectAll()
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi", '请先加载文件')

    def LoadProcess(self):
        self.clearcontext_all()
        if self.comboBoxfiletype.currentIndex() == 1:  # xls
            QMessageBox.about(self, "hi", '不支持 xls 格式文件')
        elif self.comboBoxfiletype.currentIndex() == 0:  # xlsx
            items = self.listWidget.selectedItems()
            if len(items) == 0:
                QMessageBox.about(self, "hi", '请先选择文件')
                print("YES")
            else:
                self.infos = {}
                for i in list(items):
                    file_path = str(i.text())
                    wb = load_workbook(filename=file_path)
                    print("YES")
                    name = os.path.split(file_path)[-1]

                    sheet_names = wb.sheetnames

                    sheets_dict = {}
                    for s in sheet_names:
                        sheets_dict[s] = []
                    self.infos[name] = {'path': file_path, 'sheet_names': sheets_dict}
                    print("NO")
                    wb.close()
                for k in self.infos.keys():
                    self.comboBox_wb.addItem(k)
                k = self.comboBox_wb.itemText(0)
                sheets = list(self.infos[k]['sheet_names'].keys())
                print("YES")
                for s in sheets:
                    self.comboBox_ws.addItem(s)
                self.activate_file[0] = self.infos[k]['path']
                self.activate_file[1] = list(self.infos[k]['sheet_names'].keys())[0]
                print("此处为错误")
                self.show_excel()
        print('可以预览文件')

    def first_function(self):
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi", '请先选择文件')
        else:
            self.infos = {}
            for i in list(items):
                file_path = str(i.text())
                if file_path.split('.')[-1] == 'xlsx':
                    founction_1(file_path)
                    print("功能一成功！！")
                    print("文件保存在result文件夹内,文件名为first_function.xlsx，如果已有该文件，则会将其覆盖")
                else:
                    QMessageBox.about(self, "hi", '不支持xlsx以外的文件')

    def second_function(self):
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi", '请先选择文件')
        else:
            self.infos = {}
            for i in list(items):
                file_path = str(i.text())
                if file_path.split('.')[-1] == 'xlsx':
                    handle_excel_2(file_path)
                    print("功能二已完成！")
                    print("文件保存在result文件夹内,文件名为second_function.xlsx，如果已有该文件，则会将其覆盖")
                else:
                    QMessageBox.about(self, "hi", '不支持xlsx以外的文件')

    def third_function(self):
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi", '请先选择文件')
        else:
            self.infos = {}
            for i in list(items):
                file_path = str(i.text())
                if file_path.split('.')[-1] == 'xlsx':
                    founction_3(file_path)
                    print('功能三已完成！')
                    print("文件保存在result文件夹内,文件名为third_function.xlsx，如果已有该文件，则会将其覆盖")
                else:
                    QMessageBox.about(self, "hi", '不支持xlsx以外的文件')

    def forth_function(self):
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi", '请先选择文件')
        else:
            self.infos = {}
            for i in list(items):
                file_path = str(i.text())
                if file_path.split('.')[-1] == 'xlsx':
                    list_x, list_y = excel_list_read_function_4(file_path)
                    mpl.rcParams['font.sans-serif'] = ['STZhongsong']  # 指定默认字体，解决plot不能显示中文问题
                    mpl.rcParams['axes.unicode_minus'] = False

                    # plt.figure(dpi=300,figsize=(24,8))

                    # plt.figure(dpi=105,facecolor='red')

                    plt.plot(list_x, list_y, color="red", linestyle="solid", linewidth=1.5, marker="*", mec='r',
                             mfc='w', markersize=12, label="店铺销售趋势")

                    '''
                    color 控制线的颜色
                    linestyle 控制线的风格 solid:实线
                    linewidth 控制线的粗细
                    markersize 控制标记大小
                    '''
                    plt.title("电力曲线图", loc="center")

                    for a, b in zip(list_x, list_y):
                        plt.text(a, b, b, ha="center", va="bottom", fontsize=15)
                    plt.grid(True)  # 显示网格线
                    plt.xlabel('时间', fontsize=10, color='red', fontweight='bold', loc='center',
                               labelpad=-4)
                    '''
                    xlabel 显示横坐标标题
                    fontsize 设置字体大小
                    loc 设置标签位置
                    labelpad 与轴的距离
                    '''
                    plt.ylabel("用电量", color="red")  # 显示纵坐标标题
                    plt.legend(loc=2)  # 显示图例
                    '''
                    loc 可以通过设置loc的参数来调整图例的位置
                    0 自动选择最合适的位置,1 右上角,2 左上角,3 左下角,4 右下角,5 右侧
                    6 左侧中心位置,7 右侧中心位置,8 底部中心位置,9 顶部中心位置,10 正中心位置 
                    '''
                    # 设置坐标轴的刻度
                    plt.xticks(rotation=20)  # 设置rotation X轴标题的倾斜角度
                    # plt.yticks(np.arange(20,120,20),[20,40,60,80,100,120])

                    # 关闭坐标轴
                    # plt.axis("off")
                    # plt.show()
                    if os.path.exists(r"../result/forth_result.png"):
                        os.remove(r"../result/forth_result.png")
                    plt.savefig(r"../result/forth_result.png")
                    print("功能四已完成！")
                    print("文件保存在result文件夹内,文件名为forth_result.png，如果已有该文件，则会将其覆盖")
                    plt.clf()
                elif file_path.split('.')[-1] == 'xls':
                    list_x, list_y = excel_list_read_function_4(file_path)
                    mpl.rcParams['font.sans-serif'] = ['STZhongsong']  # 指定默认字体，解决plot不能显示中文问题
                    mpl.rcParams['axes.unicode_minus'] = False
                    plt.rcParams['font.sans-serif'] = ['SimHei']  # 折线图中需显示汉字时，得加上这一行

                    plt.plot(list_x, list_y, color="red", marker='o', markersize=3, linestyle="solid", label="电力使用趋势")

                    '''
                    color 控制线的颜色
                    linestyle 控制线的风格 solid:实线
                    linewidth 控制线的粗细
                    markersize 控制标记大小
                    '''
                    plt.title("电力曲线图", loc="center")

                    for a, b in zip(list_x, list_y):
                        plt.text(a, b, b, ha="center", va="bottom", fontsize=10)
                    plt.grid(True)  # 显示网格线
                    plt.xlabel('时间', fontsize=10, color='red', fontweight='bold', loc='center',
                               labelpad=-4)
                    '''
                    xlabel 显示横坐标标题
                    fontsize 设置字体大小
                    loc 设置标签位置
                    labelpad 与轴的距离
                    '''
                    plt.ylabel("用电量", color="red")  # 显示纵坐标标题
                    plt.legend(loc=2)  # 显示图例
                    '''
                    loc 可以通过设置loc的参数来调整图例的位置
                    0 自动选择最合适的位置,1 右上角,2 左上角,3 左下角,4 右下角,5 右侧
                    6 左侧中心位置,7 右侧中心位置,8 底部中心位置,9 顶部中心位置,10 正中心位置
                    '''
                    # 设置坐标轴的刻度
                    plt.xticks(rotation=20)  # 设置rotation X轴标题的倾斜角度
                    # plt.yticks(np.arange(20,120,20),[20,40,60,80,100,120])

                    # 关闭坐标轴
                    # plt.axis("off")
                    # plt.show()
                    if os.path.exists(r"../result/forth_result.png"):
                        os.remove(r"../result/forth_result.png")
                    plt.savefig(r"../result/forth_result.png")
                    print("功能四已完成！")
                    print("文件保存在result文件夹内,文件名为forth_result.png，如果已有该文件，则会将其覆盖")
                    plt.clf()
                else:
                    QMessageBox.about(self, "hi", '不支持xlsx以外的文件')

    def fifth_function(self):
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi", '请先选择文件')
        else:
            self.infos = {}
            for i in list(items):
                file_path = str(i.text())
                if file_path.split('.')[-1] == 'xlsx':
                    result_list = mat(excel_list_read_function_5(file_path))
                    myCentroids, clusterAssing = k_means(result_list, 4)
                    plt.scatter(array(result_list)[:, 0], array(result_list)[:, 1], c=array(clusterAssing)[:, 0].T)
                    plt.scatter(myCentroids[:, 0].tolist(), myCentroids[:, 1].tolist(), c="r", label="四个聚类中心点")
                    plt.xlabel('缴费次数')
                    plt.ylabel('缴费总金额')
                    # 汉字字体，优先使用楷体，找不到则使用黑体
                    plt.legend(loc="upper right")
                    plt.rcParams['font.sans-serif'] = ['Kaitt', 'SimHei']

                    # 正常显示负号
                    plt.rcParams['axes.unicode_minus'] = False
                    print("成功！")
                    # plt.show()
                    if os.path.exists(r"../result/fifth_result.png"):
                        os.remove(r"../result/fifth_result.png")
                    plt.savefig(r"../result/fifth_result.png")
                    plt.clf()
                    print('功能五已完成！')
                    print("文件保存在result文件夹内,文件名为fifth_result.png，如果已有该文件，则会将其覆盖")
                else:
                    QMessageBox.about(self, "hi", '不支持xlsx以外的文件')

    def set_progressbar_value(self, value):
        self.progressBar.setValue(value)

    def set_lcdnumber_value(self, value):
        self.lcdNumber.display(value)

    def wbActivated(self, index):
        self.clearcontext_show()
        wb_k = self.comboBox_wb.itemText(index)
        sheets = list(self.infos[wb_k]['sheet_names'].keys())
        self.comboBox_ws.clear()
        for s in sheets:
            self.comboBox_ws.addItem(s)
        self.activate_file[0] = self.infos[wb_k]['path']
        self.activate_file[1] = list(self.infos[wb_k]['sheet_names'].keys())[0]
        self.show_excel()

    def wsActivated(self, index):
        ws_k = self.comboBox_ws.itemText(index)
        self.activate_file[1] = ws_k
        self.show_excel()

    def handleItemClick(self, item):
        cont = item.text()
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_r1.clear()
        row = item.row() + 1
        column = item.column() + 1
        # =======对合并的单元格取idx
        for p in self.merge_position:
            if row == p[0] and column == p[1]:
                row = row + (p[2] - p[0])
                break
        # =======对合并的单元格取idx
        self.comboBox_x.addItem(str(row))
        self.comboBox_y.addItem(str(column))
        self.comboBox_r1.addItem(str(row + 1))

    def show_excel(self):
        self.merge_position = []
        path = self.activate_file[0]
        sheetname = self.activate_file[1]
        wb = load_workbook(filename=path)
        ws = wb[sheetname]
        num_row = ws.max_row
        num_column = ws.max_column
        self.tableWidget.setColumnCount(num_column)
        self.tableWidget.setRowCount(num_row)
        print("NEX")
        # ======合并单元格=======
        merge_idx = ws.merged_cells
        merge_idx = get_merge_cell_list(merge_idx)

        for i in range(len(merge_idx)):
            m_idx = merge_idx[i]
            self.tableWidget.setSpan(m_idx[0] - 1, m_idx[1] - 1, m_idx[2] - m_idx[0] + 1, m_idx[3] - m_idx[1] + 1)
            self.merge_position.append([m_idx[0], m_idx[1], m_idx[2]])  # [x1,y1,range]
        # ======合并单元格=======

        # ======单元格大小=======
        for i in range(1, num_row + 1):
            h = ws.row_dimensions[i].height
            if h is not None:
                self.tableWidget.setRowHeight(i - 1, h)

        self.comboBox_r2.clear()
        for i in range(1, num_row + 1):
            self.comboBox_r2.addItem(str(num_row - i + 1))
            row_sizes = []
            for j in range(1, num_column + 1):
                cell = ws.cell(row=i, column=j)
                if cell.value is not None:
                    item = QTableWidgetItem(str(cell.value))
                    assign_style_qt(item, cell)
                else:
                    item = QTableWidgetItem()
                self.tableWidget.setItem(i - 1, j - 1, item)
        print("i can")

        wb.close()


app = QtWidgets.QApplication(sys.argv)
window = anaxcelhandler()
app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
window.show()
sys.exit(app.exec_())

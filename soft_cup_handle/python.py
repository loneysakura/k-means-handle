# import matlab
# import matlab.engine
#
# engine = matlab.engine.start_matlab()

import openpyxl
import xlrd
from numpy import *
from matplotlib import pyplot as plt
from openpyxl.utils import get_column_letter


def load_data_set(fileName):
    """加载数据集"""
    dataSet = []
    fr = open(fileName)
    # 打开文件 读取文件中的内容，然后将文件中的内容表示为数据集
    for line in fr.readlines():
        # 按tab分割字符，将每行元素分割为list的元素
        curLine = line.strip().split('\t')
        # 用list函数把map函数返回的迭代器遍历展开成一个列表
        # print(curLine)
        # 其中map(float, curLine)表示把列表的每个值用float函数转成float型，并返回迭代器
        fltLine = list(map(float, curLine))
        dataSet.append(fltLine)
    # print(dataSet)
    return dataSet


def distance_euclidean(vector1, vector2):
    """计算欧式距离"""
    return sqrt(sum(power(vector1 - vector2, 2)))


def rand_center(dataSet, k):
    """构建一个包含K个随机质心的集合"""
    n = shape(dataSet)[1]

    # 初始化质心，创建(k,n)个以0填充的矩阵
    centroids = mat(zeros((k, n)))
    # 遍历特征值
    for j in range(n):
        # 计算每一列的最小值
        # print(dataSet[:, j])
        minJ = min(dataSet[:, j])
        # 计算每一列的范围值
        rangeJ = float(max(dataSet[:, j]) - minJ)
        # 计算每一列的质心，并将其赋值给centroids
        centroids[:, j] = minJ + rangeJ * random.rand(k, 1)
    return centroids  # 返回质心


def k_means(dataSet, k, distMeas=distance_euclidean, creatCent=rand_center):
    """K-means聚类算法"""
    m = shape(dataSet)[0]  # 行数
    # 建立簇分配结果矩阵，第一列存放该数据所属中心点，第二列是该数据到中心点的距离
    clusterAssment = mat(zeros((m, 2)))
    centroids = creatCent(dataSet, k)  # 质心，即聚类点
    # 用来判定聚类是否收敛
    clusterChanged = True
    while clusterChanged:
        clusterChanged = False
        for i in range(m):  # 把每一个数据划分到离他最近的中心点
            minDist = inf  # 无穷大
            minIndex = -1  # 初始化
            for j in range(k):
                # 计算各点与新的聚类中心的距离
                distJI = distMeas(centroids[j, :], dataSet[i, :])
                if distJI < minDist:
                    # 如果第i个数据点到第J中心点更近，则将i归属为j
                    minDist = distJI
                    minIndex = j
            # 如果分配发生变化，则需要继续迭代
            if clusterAssment[i, 0] != minIndex:
                clusterChanged = True
            # 并将第i个数据点的分配情况存入字典
            clusterAssment[i, :] = minIndex, minDist ** 2
        # print(centroids)
        for cent in range(k):  # 重新计算中心点
            # 去第一列等于cent的所有列
            ptsInClust = dataSet[nonzero(clusterAssment[:, 0].A == cent)[0]]
            # 算出这些数据的中心点
            centroids[cent, :] = mean(ptsInClust, axis=0)
    return centroids, clusterAssment


def excel_list_read_function_5(excel_path):
    wb = openpyxl.load_workbook(excel_path)  # 上传excel表
    ws = wb.active
    excel_max_row = 0
    list_iterator = []
    result_list = []
    temporary_list = []
    excel_max_col = ws.max_column  # excel表最大列
    for x in range(1, 1048576):
        if ws['A' + str(x)].value is None and ws['B' + str(x)].value is None:
            excel_max_row = x
            break
    # excel_max_row = ws.max_row  # excel表最大行

    for iterator_col in range(1, excel_max_col):
        col = get_column_letter(iterator_col + 1)
        result = ws[get_column_letter(iterator_col + 1) + str(1)].value
        if '缴费次数' in result:
            list_iterator.append(col)
            money_last = col
        elif '缴费总金额' in result:
            list_iterator.append(col)

    for iterator_row in range(2, excel_max_row):
        temporary_list = []
        for iterator_col in list_iterator:
            # print(ws[iterator_col + str(iterator_row)])
            temporary_list.append(ws[iterator_col + str(iterator_row)].value)
        result_list.append(temporary_list)

    return result_list


def excel_list_read_function_4(excel_path):
    x = 0
    y = 0
    result_list_x = []
    result_list_y = []
    # 打开文件
    workbook = xlrd.open_workbook(excel_path)
    # 读取sheet页
    sheet = workbook.sheet_by_index(0)
    # 获取表的行列数
    rows = sheet.nrows
    cols = sheet.ncols
    # 获取表中数值
    for col in range(0, cols):
        result = sheet.cell(0, col).value
        if '小时' in result:
            x = col
        elif '电量' in result:
            y = col

    for row in range(1, rows-2):
        result_list_x.append(sheet.cell(row, x).value)
        result_list_y.append(sheet.cell(row, y).value)

    return result_list_x,result_list_y

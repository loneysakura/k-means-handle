import os.path
import re
from typing import List, Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import numpy as np

from function_1_row_if import if_function_1


def inner_function_1_extra(excel_path):
    wb = openpyxl.load_workbook(excel_path)  # 上传excel表
    ws = wb.active

    list_id = []
    list_times = []
    list_money = []
    list_iterator = []
    money = 0
    times = 0
    money_last = ''
    person_money_list = []
    excel_max_row = 0

    excel_max_col = ws.max_column  # excel表最大列
    excel_max_row = excel_read_max_row(ws)
    # excel_max_row = ws.max_row  # excel表最大行

    for iterator_col in range(0, excel_max_col):
        col = get_column_letter(iterator_col + 1)
        result = ws[get_column_letter(iterator_col + 1) + str(1)].value
        if '缴费金额' in result:
            list_iterator.append(col)
            money_last = col
        elif '缴费日期' in result:
            list_iterator.append(col)
        elif '编号' in result:
            list_iterator.append(col)
    # print(excel_max_row)
    for iterator_row in range(2, excel_max_row + 1):
        for iterator_col in list_iterator:
            break_num = 0
            list_id, list_times, list_money, money, times, break_num = \
                if_function_1(ws, iterator_col, iterator_row, list_id, list_times, list_money, money,
                              times, excel_max_row, money_last, break_num)
            if break_num == 1:
                break
    list_key = [[''] for y in range(len(list_iterator))]  # 创建双重空列表

    list_key[0] = list_key[0] + list_id
    list_key[1] = list_key[1] + list_money
    list_key[2] = list_key[2] + list_money

    for x in range(0, len(list_key)):
        del list_key[x][0]
    return list_key, excel_max_row


def inner_function_1(excel_path):
    wb = openpyxl.load_workbook(excel_path)  # 上传excel表
    ws = wb.active

    excel_max_row = 0
    list_id = []
    list_times = []
    list_money = []
    list_iterator = []
    money = 0
    times = 0
    money_last = ''
    person_money_list = []

    excel_max_col = ws.max_column  # excel表最大列
    excel_max_row = excel_read_max_row(ws)

    for iterator_col in range(0, excel_max_col):
        col = get_column_letter(iterator_col + 1)
        result = ws[get_column_letter(iterator_col + 1) + str(1)].value
        if '缴费金额' in result:
            list_iterator.append(col)
            money_last = col
        elif '缴费日期' in result:
            list_iterator.append(col)
        elif '编号' in result:
            list_iterator.append(col)
    # print(excel_max_row)
    # print(list_iterator)
    # print(excel_max_row)
    # print(list_iterator)
    for iterator_row in range(2, excel_max_row + 1):
        for iterator_col in list_iterator:
            break_num = 0
            list_id, list_times, list_money, money, times, break_num = \
                if_function_1(ws, iterator_col, iterator_row, list_id, list_times, list_money, money,
                              times, excel_max_row, money_last, break_num)
            if break_num == 1:
                break
    list_key = [[''] for y in range(len(list_iterator))]  # 创建双重空列表

    list_key[0] = list_key[0] + list_id
    list_key[1] = list_key[1] + list_times
    list_key[2] = list_key[2] + list_money
    print("YES")
    for x in range(0, len(list_key)):
        del list_key[x][0]
    return list_key


def founction_1(excel_path):
    list_key = inner_function_1(excel_path)
    # print(list_key)
    list_key = list(list_key)
    list_main = ['用户编号', '缴费次数', '缴费总金额']
    result_dict = dict(zip(list_main, list_key))
    df = pd.DataFrame(result_dict)
    total_df = df
    if os.path.exists(r'./result/first_function.xlsx'):
        os.remove(r'./result/first_function.xlsx')
    total_df.to_excel(r'../result/first_function.xlsx')


# founction_1(r'D:\Desktop\user_function_1.xlsx')


def handle_excel_2(excel_path):
    def test(now_money, now_times, avg_money, avg_times):
        if now_money >= avg_money and now_times >= avg_times:
            return '高价值型客户'
        elif now_money >= avg_money and now_times < avg_times:
            return '大众型客户'

        elif now_money < avg_money and now_times >= avg_times:
            return '潜力型客户'

        else:
            return '低价值型客户'

    wb = openpyxl.load_workbook(excel_path)  # 上传excel表
    ws = wb.active

    excel_max_row = 0

    excel_max_col = ws.max_column  # excel表最大列
    excel_max_row = excel_read_max_row(ws)
    # excel_max_row = ws.max_row  # excel表最大行

    list_iterator = []

    total_money = 0
    total_times = 0
    money_last = ''
    times_last = ''
    num_last = ''
    result_list = []

    for iterator_col in range(1, excel_max_col):
        col = get_column_letter(iterator_col + 1)
        result = ws[get_column_letter(iterator_col + 1) + str(1)].value
        if '缴费次数' in result:
            list_iterator.append(col)
            times_last = col
        elif '缴费总金额' in result:
            list_iterator.append(col)
            money_last = col
        elif '编号' in result:
            list_iterator.append(col)
            num_last = col

    for iterator_row in range(2, excel_max_row):
        total_money += ws[money_last + str(iterator_row)].value
        total_times += ws[times_last + str(iterator_row)].value
    total_num = excel_max_row - 1
    avg_times = total_times / total_num
    avg_money = total_money / total_num

    total_df = pd.DataFrame(pd.read_excel(excel_path))
    if 'Unnamed: 0' in total_df.columns:
        total_df = total_df.drop(labels='Unnamed: 0', axis=1)

    for iterator_row in range(2, excel_max_row + 1):
        now_money = ws[money_last + str(iterator_row)].value
        now_times = ws[times_last + str(iterator_row)].value
        if now_money >= avg_money and now_times >= avg_times:
            result_list.append('高价值型客户')
        elif now_money >= avg_money and now_times < avg_times:
            result_list.append('大众型客户')
        elif now_money < avg_money and now_times >= avg_times:
            result_list.append('潜力型客户')
        else:
            result_list.append('低价值型客户')

    size = len(result_list)

    total_df['客户类型'] = result_list

    # print(total_df)
    # total_df.to_excel('../result/second_function.xlsx')
    if os.path.exists(r'../result/second_function.xlsx'):
        os.remove(r'../result/second_function.xlsx')
    total_df.to_excel(r'../result/second_function.xlsx')


def founction_3(excel_path):
    result_3_list = []
    total_money = 0
    total_times = 0
    list_top_value = []
    result_value = []
    result_key = []

    list_key, excel_max_row = inner_function_1_extra(excel_path)

    total_num = excel_max_row - 1
    for iterator_row in range(0, len(list_key[0])):
        total_money += list_key[2][iterator_row]
        total_times += list_key[1][iterator_row]

    for iterator_list_key in range(0, len(list_key[0])):
        person_times = list_key[1][iterator_list_key]
        avg_times = total_times / total_num
        person_avg_money = float(list_key[2][iterator_list_key]) / float(list_key[1][iterator_list_key])
        result_3_list.append(round((float(person_times) - avg_times) * person_avg_money, 2))

    key_dict = dict(zip(result_3_list, list_key[0]))
    result_3_list.sort()
    result_3_list.reverse()

    for top_5 in range(0, 5):
        list_top_value.append(key_dict[result_3_list[top_5]])

    result_key = ['用户编号', '加权价值']
    list_value = [[''] for y in range(2)]  # 创建双重空列表
    list_value[0] += list_top_value
    list_value[1] += result_3_list[:5]

    for x in range(0, len(list_value)):
        del list_value[x][0]
    result_dict = dict(zip(result_key, list_value))

    df = pd.DataFrame(result_dict)
    total_df = df
    if os.path.exists(r'../result/third_function.xlsx'):
        os.remove(r'../result/third_function.xlsx')
    total_df.to_excel(r'../result/third_function.xlsx')


def excel_read_max_row(ws):
    for x in range(1, 1048576):
        if ws['A' + str(x)].value is None and ws['B' + str(x)].value is None:
            excel_max_row = x - 1
            break
    return excel_max_row
